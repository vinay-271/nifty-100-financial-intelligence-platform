import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


PEER_GROUPS = [
    "Automobiles",
    "Consumer Finance",
    "FMCG",
    "IT Services",
    "Life Insurance",
    "Oil & Gas",
    "Pharmaceuticals",
    "Power & Utilities",
    "Private Banks",
    "Public Sector Banks",
    "Steel",
]

LOW_FILL = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid",
)

MID_FILL = PatternFill(
    start_color="FFEB9C",
    end_color="FFEB9C",
    fill_type="solid",
)

HIGH_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid",
)


def percentile_fill(value):
    """Return the appropriate fill for a percentile rank."""
    if pd.isna(value):
        return None

    if value < 0.33:
        return LOW_FILL

    if value < 0.67:
        return MID_FILL

    return HIGH_FILL


def auto_width(ws, min_width=10, max_width=25):
    """Automatically size worksheet columns."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, min_width),
            max_width,
        )


def export_peer_comparison(
    db_path="db/nifty100.db",
    output_path="output/peer_comparison.xlsx",
):
    """
    Export peer percentile rankings from SQLite
    into an Excel workbook with one sheet per peer group.
    """

    db_path = Path(db_path)
    output_path = Path(output_path)

    connection = sqlite3.connect(db_path)

    try:
        query = """
            SELECT
                company_id,
                peer_group_name,
                metric,
                value,
                percentile_rank,
                year
            FROM peer_percentiles
            ORDER BY
                peer_group_name,
                metric,
                percentile_rank DESC
        """

        df = pd.read_sql(query, connection)

    finally:
        connection.close()

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    columns = [
        "company_id",
        "metric",
        "value",
        "percentile_rank",
        "year",
    ]

    for peer_group in PEER_GROUPS:
        sheet_df = df[
            df["peer_group_name"] == peer_group
        ].copy()

        sheet_df = sheet_df[columns]

        ws = workbook.create_sheet(
            title=peer_group[:31]
        )

        # Header
        for column_index, column_name in enumerate(
            columns,
            start=1,
        ):
            cell = ws.cell(
                row=1,
                column=column_index,
                value=column_name,
            )

            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                wrap_text=True,
            )

        # Data
        for row_index, (_, row) in enumerate(
            sheet_df.iterrows(),
            start=2,
        ):
            for column_index, column_name in enumerate(
                columns,
                start=1,
            ):
                value = row[column_name]

                cell = ws.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

                if column_name == "value":
                    cell.number_format = "0.00"

                elif column_name == "percentile_rank":
                    cell.number_format = "0.00%"

                    fill = percentile_fill(value)

                    if fill:
                        cell.fill = fill

        # Excel usability
        last_row = len(sheet_df) + 1

        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}{last_row}"
        )

        ws.freeze_panes = "A2"

        auto_width(ws)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(output_path)

    return output_path


if __name__ == "__main__":
    path = export_peer_comparison()

    print(
        f"Peer comparison workbook generated: {path}"
    )
