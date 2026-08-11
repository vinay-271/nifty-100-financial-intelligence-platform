import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
import pandas as pd


GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

FORMATS = {
    "pct": "0.00%",
    "ratio": "0.00",
    "currency_cr": "#,##0.00",
    "score": "0.0",
}

KPI_COLUMNS = [
    "company_id", "company_name", "sector", "year", "composite_quality_score",
    "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
    "debt_to_equity", "interest_coverage", "free_cash_flow_cr",
    "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "operating_profit_margin_pct", "asset_turnover",
    "market_cap_cr", "pe_ratio", "pb_ratio", "dividend_yield_pct",
]

COLUMN_FORMATS = {
    "return_on_equity_pct": "pct",
    "return_on_capital_employed_pct": "pct",
    "net_profit_margin_pct": "pct",
    "debt_to_equity": "ratio",
    "interest_coverage": "ratio",
    "free_cash_flow_cr": "currency_cr",
    "revenue_cagr_5yr": "pct",
    "pat_cagr_5yr": "pct",
    "eps_cagr_5yr": "pct",
    "operating_profit_margin_pct": "pct",
    "asset_turnover": "ratio",
    "market_cap_cr": "currency_cr",
    "pe_ratio": "ratio",
    "pb_ratio": "ratio",
    "dividend_yield_pct": "pct",
    "composite_quality_score": "score",
}

PRESET_THRESHOLDS = {
    "quality_compounder": {
        "return_on_equity_pct": (15, ">="),
        "debt_to_equity": (1.0, "<="),
        "free_cash_flow_cr": (0, ">="),
        "revenue_cagr_5yr": (10, ">="),
    },
    "value_pick": {
        "pe_ratio": (30, "<="),
        "pb_ratio": (5.0, "<="),
        "debt_to_equity": (2.0, "<="),
        "dividend_yield_pct": (1.0, ">="),
    },
    "growth_accelerator": {
        "pat_cagr_5yr": (20, ">="),
        "revenue_cagr_5yr": (15, ">="),
        "debt_to_equity": (2.0, "<="),
    },
    "dividend_champion": {
        "dividend_yield_pct": (2.0, ">="),
        "dividend_payout": (80, "<="),
        "free_cash_flow_cr": (0, ">="),
    },
    "debt_free_blue_chip": {
        "debt_to_equity": (0, "=="),
        "return_on_equity_pct": (12, ">="),
        "sales": (5000, ">="),
    },
    "turnaround_watch": {
        "revenue_cagr_3yr": (10, ">="),
        "free_cash_flow_cr": (0, ">="),
        "debt_to_equity_declining": (True, "=="),
    },
}

PRESET_SHEET_NAMES = {
    "quality_compounder": "Quality Compounder",
    "value_pick": "Value Pick",
    "growth_accelerator": "Growth Accelerator",
    "dividend_champion": "Dividend Champion",
    "debt_free_blue_chip": "Debt Free Blue Chip",
    "turnaround_watch": "Turnaround Watch",
}

PRESET_ORDER = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
    "turnaround_watch",
]


def check_threshold(value, threshold, operator):
    if pd.isna(value):
        return False
    if operator in (">=", "min"):
        return value >= threshold
    if operator in ("<=", "max"):
        return value <= threshold
    if operator in ("==", "eq"):
        return value == threshold
    return False


def auto_width(ws, min_width=10, max_width=25):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def export_screener_output(engine, df: pd.DataFrame, output_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for preset_name in PRESET_ORDER:
        result = engine.run_preset(preset_name)

        for col in KPI_COLUMNS:
            if col not in result.columns:
                result[col] = None

        sheet_df = result[KPI_COLUMNS].copy()

        sheet_name = PRESET_SHEET_NAMES[preset_name]
        ws = wb.create_sheet(title=sheet_name[:31])

        for col_idx, col_name in enumerate(KPI_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        thresholds = PRESET_THRESHOLDS.get(preset_name, {})

        for row_idx, (_, row) in enumerate(sheet_df.iterrows(), 2):
            for col_idx, col_name in enumerate(KPI_COLUMNS, 1):
                value = row[col_name]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                fmt = COLUMN_FORMATS.get(col_name)
                if fmt:
                    cell.number_format = FORMATS[fmt]

                if col_idx >= 6 and col_name in thresholds:
                    thresh_val, thresh_op = thresholds[col_name]
                    if check_threshold(value, thresh_val, thresh_op):
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL

        ws.auto_filter.ref = f"A1:{get_column_letter(len(KPI_COLUMNS))}{len(sheet_df) + 1}"
        ws.freeze_panes = "A2"
        auto_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)