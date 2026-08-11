import pytest
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

from src.screener.exporter import (
    export_screener_output,
    KPI_COLUMNS,
    PRESET_THRESHOLDS,
    PRESET_SHEET_NAMES,
    PRESET_ORDER,
    check_threshold,
    GREEN_FILL,
    RED_FILL,
    FORMATS,
    COLUMN_FORMATS,
)


class MockEngine:
    def __init__(self, df):
        self._df = df

    def run_preset(self, name):
        result = self._df.copy()
        if "composite_quality_score" in result.columns:
            result = result.sort_values(
                by="composite_quality_score",
                ascending=False,
                na_position="last",
            ).reset_index(drop=True)
        return result


@pytest.fixture
def sample_data():
    return pd.DataFrame({
        "company_id": ["TCS", "INFY", "RELIANCE", "HDFCBANK"],
        "company_name": ["Tata Consultancy", "Infosys", "Reliance Industries", "HDFC Bank"],
        "sector": ["IT Services", "IT Services", "Oil & Gas", "Financials"],
        "year": ["Mar 2024"] * 4,
        "composite_quality_score": [85.5, 72.3, 45.1, 60.0],
        "return_on_equity_pct": [45.0, 30.0, 12.0, 18.0],
        "return_on_capital_employed_pct": [50.0, 35.0, 15.0, 20.0],
        "net_profit_margin_pct": [20.0, 18.0, 10.0, 25.0],
        "debt_to_equity": [0.1, 0.2, 1.5, 0.8],
        "interest_coverage": [100.0, 50.0, 5.0, 25.0],
        "free_cash_flow_cr": [40000, 20000, 5000, 15000],
        "revenue_cagr_5yr": [12.0, 10.0, 8.0, 15.0],
        "pat_cagr_5yr": [15.0, 12.0, 5.0, 20.0],
        "eps_cagr_5yr": [14.0, 11.0, 6.0, 18.0],
        "operating_profit_margin_pct": [25.0, 22.0, 15.0, 30.0],
        "asset_turnover": [1.5, 1.2, 0.8, 1.0],
        "market_cap_cr": [1400000, 600000, 1800000, 900000],
        "pe_ratio": [28.0, 25.0, 15.0, 20.0],
        "pb_ratio": [12.0, 8.0, 2.0, 3.0],
        "dividend_yield_pct": [1.5, 2.0, 0.5, 1.0],
        "sales": [250000, 150000, 800000, 200000],
        "dividend_payout": [40.0, 50.0, 30.0, 45.0],
        "revenue_cagr_3yr": [11.0, 9.0, 7.0, 14.0],
        "debt_to_equity_declining": [True, True, False, True],
    })


def test_workbook_creation(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    assert output_path.exists()


def test_six_sheets(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    assert len(wb.sheetnames) == 6
    expected = [PRESET_SHEET_NAMES[p] for p in PRESET_ORDER]
    assert wb.sheetnames == expected


def test_required_columns(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(KPI_COLUMNS) + 1)]
        assert headers == KPI_COLUMNS


def test_correct_sorting(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    ws = wb["Quality Compounder"]
    scores = [ws.cell(row=r, column=KPI_COLUMNS.index("composite_quality_score") + 1).value
              for r in range(2, ws.max_row + 1)]
    assert scores == sorted(scores, reverse=True)


def test_number_formats(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    ws = wb["Quality Compounder"]
    for col_name, fmt_key in COLUMN_FORMATS.items():
        col_idx = KPI_COLUMNS.index(col_name) + 1
        cell = ws.cell(row=2, column=col_idx)
        expected_fmt = FORMATS[fmt_key]
        assert cell.number_format == expected_fmt, f"{col_name}: got {cell.number_format}, expected {expected_fmt}"


def test_color_coding_green(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    ws = wb["Quality Compounder"]
    # TCS ROE=45 >= 15 -> green
    col_idx = KPI_COLUMNS.index("return_on_equity_pct") + 1
    cell = ws.cell(row=2, column=col_idx)
    assert cell.fill.start_color.rgb == "00C6EFCE"


def test_color_coding_red(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    ws = wb["Quality Compounder"]
    # Find RELIANCE row (ROE=12 < 15 -> red)
    col_idx = KPI_COLUMNS.index("return_on_equity_pct") + 1
    rel_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=KPI_COLUMNS.index("company_id") + 1).value == "RELIANCE":
            rel_row = r
            break
    assert rel_row is not None
    cell = ws.cell(row=rel_row, column=col_idx)
    assert cell.fill.start_color.rgb == "00FFC7CE"


def test_freeze_panes(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.freeze_panes == "A2"


def test_autofilter(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.auto_filter.ref is not None


def test_auto_width(sample_data, tmp_path):
    engine = MockEngine(sample_data)
    output_path = tmp_path / "test_output.xlsx"
    export_screener_output(engine, sample_data, output_path)
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            assert width is not None and width >= 10


def test_check_threshold():
    assert check_threshold(20, 15, ">=") is True
    assert check_threshold(10, 15, ">=") is False
    assert check_threshold(20, 30, "<=") is True
    assert check_threshold(40, 30, "<=") is False
    assert check_threshold(5, 5, "==") is True
    assert check_threshold(5, 6, "==") is False
    assert check_threshold(float("inf"), 10, ">=") is True
    assert check_threshold(float("nan"), 10, ">=") is False
    assert check_threshold(None, 10, ">=") is False


def test_empty_result_handling(tmp_path):
    empty_df = pd.DataFrame(columns=KPI_COLUMNS)
    engine = MockEngine(empty_df)
    output_path = tmp_path / "test_empty.xlsx"
    export_screener_output(engine, empty_df, output_path)
    assert output_path.exists()
    wb = load_workbook(output_path)
    assert len(wb.sheetnames) == 6
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.max_row == 1


def test_missing_columns_filled_with_none(sample_data, tmp_path):
    df = sample_data.drop(columns=["interest_coverage"])
    engine = MockEngine(df)
    output_path = tmp_path / "test_missing.xlsx"
    export_screener_output(engine, df, output_path)
    wb = load_workbook(output_path)
    ws = wb["Quality Compounder"]
    col_idx = KPI_COLUMNS.index("interest_coverage") + 1
    cell = ws.cell(row=2, column=col_idx)
    assert cell.value is None